"""
openpyxl / xlrd 기반 XLS/XLSX → Markdown 변환기
"""
from pathlib import Path
from typing import Optional

from .marker_converter import ConversionResult


def _sheet_to_markdown(ws) -> str:
    """
    워크시트를 Markdown 테이블 문자열로 변환

    Args:
        ws: openpyxl Worksheet 객체

    Returns:
        Markdown 테이블 문자열
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    # 빈 행/열 제거: 실제 데이터가 있는 범위만 추출
    # 뒤쪽 빈 열 제거
    max_col = 0
    for row in rows:
        for ci in range(len(row) - 1, -1, -1):
            if row[ci] is not None:
                max_col = max(max_col, ci + 1)
                break

    if max_col == 0:
        return ""

    # 뒤쪽 빈 행 제거
    last_row = 0
    for ri in range(len(rows) - 1, -1, -1):
        if any(c is not None for c in rows[ri][:max_col]):
            last_row = ri + 1
            break

    if last_row == 0:
        return ""

    rows = [row[:max_col] for row in rows[:last_row]]

    # Markdown 테이블 생성
    lines = []

    def cell_str(v):
        if v is None:
            return ""
        return str(v).replace("|", "\\|").replace("\n", " ")

    # 헤더 (첫 번째 행)
    header = "| " + " | ".join(cell_str(c) for c in rows[0]) + " |"
    separator = "| " + " | ".join("---" for _ in rows[0]) + " |"
    lines.append(header)
    lines.append(separator)

    # 데이터 행
    for row in rows[1:]:
        line = "| " + " | ".join(cell_str(c) for c in row) + " |"
        lines.append(line)

    return "\n".join(lines)


def convert_excel(
    excel_path: str | Path,
    output_dir: Optional[str | Path] = None,
    save_images: bool = True,
) -> ConversionResult:
    """
    단일 XLS/XLSX 파일을 Markdown으로 변환

    Args:
        excel_path: XLS 또는 XLSX 파일 경로
        output_dir: 출력 디렉토리 (None이면 원본과 같은 위치)
        save_images: (Excel에서는 미사용, 인터페이스 호환용)

    Returns:
        ConversionResult: 변환 결과
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        return ConversionResult(
            success=False,
            markdown="",
            images={},
            metadata={},
            error=f"파일을 찾을 수 없습니다: {excel_path}",
        )

    if output_dir is None:
        output_dir = excel_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    try:
        ext = excel_path.suffix.lower()

        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(str(excel_path))
            sheet_names = wb.sheet_names()
        else:
            return ConversionResult(
                success=False,
                markdown="",
                images={},
                metadata={},
                error=f"지원하지 않는 형식입니다: {ext}",
            )

        md_parts = []

        for sheet_name in sheet_names:
            if ext == ".xlsx":
                ws = wb[sheet_name]
                table_md = _sheet_to_markdown(ws)
            else:
                # xlrd 시트를 openpyxl 호환 형태로 변환
                xs = wb.sheet_by_name(sheet_name)
                rows = []
                for r in range(xs.nrows):
                    rows.append(tuple(xs.cell_value(r, c) for c in range(xs.ncols)))
                table_md = _rows_to_markdown(rows)

            if table_md:
                if len(sheet_names) > 1:
                    md_parts.append(f"## {sheet_name}\n\n{table_md}")
                else:
                    md_parts.append(table_md)

        if ext == ".xlsx":
            wb.close()

        markdown_text = "\n\n".join(md_parts) + "\n" if md_parts else ""

        # 마크다운 파일 저장
        md_filename = excel_path.stem + ".md"
        md_path = output_dir / md_filename
        md_path.write_text(markdown_text, encoding="utf-8")

        return ConversionResult(
            success=True,
            markdown=markdown_text,
            images={},
            metadata={"sheets": len(sheet_names)},
        )

    except Exception as e:
        return ConversionResult(
            success=False,
            markdown="",
            images={},
            metadata={},
            error=str(e),
        )


def _rows_to_markdown(rows: list[tuple]) -> str:
    """
    행 데이터 리스트를 Markdown 테이블로 변환 (xlrd용)

    Args:
        rows: [(val1, val2, ...), ...] 형태의 행 데이터

    Returns:
        Markdown 테이블 문자열
    """
    if not rows:
        return ""

    # 빈 열 제거
    max_col = 0
    for row in rows:
        for ci in range(len(row) - 1, -1, -1):
            v = row[ci]
            if v is not None and v != "":
                max_col = max(max_col, ci + 1)
                break

    if max_col == 0:
        return ""

    # 빈 행 제거
    last_row = 0
    for ri in range(len(rows) - 1, -1, -1):
        if any(c is not None and c != "" for c in rows[ri][:max_col]):
            last_row = ri + 1
            break

    if last_row == 0:
        return ""

    rows = [row[:max_col] for row in rows[:last_row]]

    lines = []

    def cell_str(v):
        if v is None or v == "":
            return ""
        return str(v).replace("|", "\\|").replace("\n", " ")

    header = "| " + " | ".join(cell_str(c) for c in rows[0]) + " |"
    separator = "| " + " | ".join("---" for _ in rows[0]) + " |"
    lines.append(header)
    lines.append(separator)

    for row in rows[1:]:
        line = "| " + " | ".join(cell_str(c) for c in row) + " |"
        lines.append(line)

    return "\n".join(lines)


def convert_excel_batch(
    input_dir: str | Path,
    output_dir: Optional[str | Path] = None,
    recursive: bool = False,
    save_images: bool = True,
) -> list[tuple[Path, ConversionResult]]:
    """
    폴더 내 모든 XLS/XLSX 파일을 Markdown으로 변환

    Args:
        input_dir: 입력 디렉토리
        output_dir: 출력 디렉토리 (None이면 입력과 같은 위치)
        recursive: 하위 폴더 포함 여부
        save_images: (Excel에서는 미사용, 인터페이스 호환용)

    Returns:
        list[tuple[Path, ConversionResult]]: (파일경로, 변환결과) 리스트
    """
    input_dir = Path(input_dir)

    if output_dir is None:
        output_dir = input_dir
    else:
        output_dir = Path(output_dir)

    # XLS/XLSX 파일 목록 수집
    excel_files = []
    for ext in ("*.xls", "*.xlsx"):
        if recursive:
            excel_files.extend(input_dir.rglob(ext))
        else:
            excel_files.extend(input_dir.glob(ext))

    results = []

    for excel_path in excel_files:
        # 하위 폴더 구조 유지
        if recursive:
            relative_path = excel_path.parent.relative_to(input_dir)
            current_output_dir = output_dir / relative_path
        else:
            current_output_dir = output_dir

        result = convert_excel(excel_path, current_output_dir, save_images)
        results.append((excel_path, result))

    return results
